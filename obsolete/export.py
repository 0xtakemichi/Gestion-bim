import os
import csv

def export_text_to_csv(observaciones):
    # Define la carpeta donde se guardará el archivo CSV
    csv_folder = "csv_files"
    # Comprueba si la carpeta del archivo CSV existe, si no existe, la crea
    if not os.path.exists(csv_folder):
        os.makedirs(csv_folder)
    # Construye el nombre del archivo CSV
    csv_filename = "observaciones.csv"
    csv_path = os.path.join(csv_folder, csv_filename)

    with open(csv_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        # Titulos
        writer.writerow(["Label", "Title", "Description", "Type", "Priority", "Status", "Completion","Due date", "Assigne(s)","Tags","Created by","Created on", "Last modified by","Last modified on"])

        for enum, texto in observaciones:
            writer.writerow([enum, "" , texto, "", "", "", "", "", "", "", "", "", "", ""])




import openpyxl
def export_text_to_excel(observaciones, excel_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    for index, (enum, texto) in enumerate(observaciones, start=1):
        sheet.cell(row=index, column=1).value = enum
        sheet.cell(row=index, column=2).value = texto
    
    workbook.save(excel_path)


def export_text_to_txt(observaciones, txt_path):
    with open(txt_path, 'w') as txt_file:
        for enum, texto in observaciones:
            txt_file.write(f"{enum} {texto}\n")


#txt_observaciones = "observaciones.txt"
#xlsx_observaciones = "observaciones.xlsx"

#export_text_to_excel(observaciones,xlsx_observaciones)
#export_text_to_txt(observaciones, txt_observaciones)
